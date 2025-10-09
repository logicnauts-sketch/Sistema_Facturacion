from flask import Blueprint, render_template, jsonify, request, send_file
from conexion import conectar
import math
from utils import login_required, solo_admin_required
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime

bp = Blueprint('productos', __name__)

def calcular_estado_stock(stock_actual, stock_minimo):
    if stock_actual <= stock_minimo:
        return {'clase': 'status-danger', 'texto': 'Muy bajo'}
    elif stock_actual <= stock_minimo * 1.5:
        return {'clase': 'status-warning', 'texto': 'Bajo stock'}
    return {'clase': 'status-normal', 'texto': 'Normal'}

def validar_producto(data):
    errores = []
    campos_requeridos = ['code', 'name', 'category']
    campos_numericos = ['purchase_price', 'sale_price', 'initial_quantity', 
                        'min_stock', 'max_stock', 'tax_percentage']
    
    # Validar campos requeridos
    for campo in campos_requeridos:
        if not data.get(campo) or not str(data[campo]).strip():
            nombre_campo = {'code': 'código', 'name': 'nombre', 'category': 'categoría'}[campo]
            errores.append(f"El campo {nombre_campo} es obligatorio")
    
    # Validar campos numéricos
    for campo in campos_numericos:
        try:
            valor = float(data.get(campo, 0))
            if valor < 0:
                errores.append(f"El campo {campo} no puede ser negativo")
        except (TypeError, ValueError):
            errores.append(f"El campo {campo} debe ser un número válido")
    
    # Validar stock mínimo/máximo
    try:
        min_stock = int(data.get('min_stock', 0))
        max_stock = int(data.get('max_stock', 0))
        if min_stock > max_stock:
            errores.append("El stock mínimo no puede ser mayor que el stock máximo")
    except:
        pass
    
    return errores

def obtener_productos(conn, categoria_filtro, buscar_filtro):
    cursor = conn.cursor(dictionary=True)
    query = """
        SELECT p.id, p.codigo, p.nombre, c.nombre as categoria, 
               p.precio_compra, p.precio_venta, p.stock_actual, 
               p.stock_minimo, p.stock_maximo, p.descripcion, p.impuesto
        FROM productos p
        JOIN categorias c ON p.categoria_id = c.id
    """
    condiciones, parametros = [], []
    
    if categoria_filtro and categoria_filtro != 'Todos':
        condiciones.append("c.nombre = %s")
        parametros.append(categoria_filtro)
    
    if buscar_filtro:
        condiciones.append("(p.nombre LIKE %s OR p.codigo LIKE %s OR c.nombre LIKE %s OR p.descripcion LIKE %s)")
        parametros.extend([f"%{buscar_filtro}%"] * 4)
    
    if condiciones:
        query += " WHERE " + " AND ".join(condiciones)
    
    cursor.execute(query, parametros)
    productos = cursor.fetchall()
    
    for producto in productos:
        estado = calcular_estado_stock(producto['stock_actual'], producto['stock_minimo'])
        producto['estado_clase'] = estado['clase']
        producto['estado_texto'] = estado['texto']
        producto['stock_porcentaje'] = min(100, (producto['stock_actual'] / producto['stock_maximo'] * 100) if producto['stock_maximo'] > 0 else 0)

    cursor.close()
    return productos

def manejar_db(operacion, *args, **kwargs):
    conn = conectar()
    if not conn:
        return None, "Error de conexión a la base de datos"
    
    try:
        return operacion(conn, *args, **kwargs), None
    except Exception as e:
        return None, str(e)
    finally:
        conn.close()

@bp.route('/productos')
@login_required
@solo_admin_required
def productos_page():
    categoria_filtro = request.args.get('categoria', 'Todos')
    buscar_filtro = request.args.get('buscar', '').strip()
    
    def operacion(conn):
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, nombre FROM categorias")
        categorias = cursor.fetchall()
        
        # Obtener contadores de productos por categoría
        cursor.execute("""
            SELECT c.nombre AS categoria, COUNT(p.id) AS total
            FROM categorias c
            LEFT JOIN productos p ON c.id = p.categoria_id
            GROUP BY c.nombre
        """)
        contadores = cursor.fetchall()
        contador_dict = {item['categoria']: item['total'] for item in contadores}
        
        # Contador para "Todos"
        cursor.execute("SELECT COUNT(*) AS total FROM productos")
        total_todos = cursor.fetchone()['total']
        contador_dict['Todos'] = total_todos
        
        cursor.close()
        
        productos = obtener_productos(conn, categoria_filtro, buscar_filtro)
        return {
            "categorias": categorias,
            "productos": productos,
            "categoria_activa": categoria_filtro,
            "termino_busqueda": buscar_filtro,
            "sin_productos": len(productos) == 0,
            "contadores": contador_dict
        }
    
    resultado, error = manejar_db(operacion)
    if error:
        return render_template("error.html", error=f"Error al cargar productos: {error}")
    return render_template("productos.html", **resultado)

@bp.route('/api/productos', methods=['GET', 'POST'])
def handle_productos():
    if request.method == 'GET':
        def operacion(conn):
            cursor = conn.cursor(dictionary=True)
            category = request.args.get('category')
            
            if category and category != 'Todos':
                cursor.execute("""
                    SELECT p.id, p.codigo, p.nombre, c.nombre as categoria, 
                           p.precio_compra, p.precio_venta, p.stock_actual, 
                           p.stock_minimo, p.stock_maximo, p.descripcion, p.impuesto
                    FROM productos p
                    JOIN categorias c ON p.categoria_id = c.id
                    WHERE c.nombre = %s
                """, (category,))
            else:
                cursor.execute("""
                    SELECT p.id, p.codigo, p.nombre, c.nombre as categoria, 
                           p.precio_compra, p.precio_venta, p.stock_actual, 
                           p.stock_minimo, p.stock_maximo, p.descripcion, p.impuesto
                    FROM productos p
                    JOIN categorias c ON p.categoria_id = c.id
                """)
            
            productos = cursor.fetchall()
            for p in productos:
                p['codigo'] = p['codigo'] or f"PRD-{str(p['id']).zfill(3)}"
                p['nombre'] = p['nombre'] or 'Producto sin nombre'
                p['categoria'] = p['categoria'] or 'Sin categoría'
                p['descripcion'] = p['descripcion'] or ''
            
            cursor.close()
            return {"table": "productos", "rows": productos}
        
        resultado, error = manejar_db(operacion)
        return jsonify(resultado) if resultado else jsonify({'error': error}), 500
    
    elif request.method == 'POST':
        data = request.get_json()
        errores = validar_producto(data)
        if errores:
            return jsonify({'error': 'Errores de validación', 'detalles': errores}), 400
        
        try:
            valores = (
                data['code'],
                data['name'],
                float(data['purchase_price']),
                float(data['sale_price']),
                int(data['initial_quantity']),
                int(data['min_stock']),
                int(data['max_stock']),
                data.get('description', ''),
                float(data['tax_percentage'])
            )
        except (TypeError, ValueError) as e:
            return jsonify({'error': f'Tipos de datos inválidos: {str(e)}'}), 400
        
        if valores[5] > valores[6]:
            return jsonify({'error': 'El stock mínimo no puede ser mayor que el stock máximo'}), 400
        
        def operacion(conn):
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT id FROM categorias WHERE nombre = %s", (data['category'],))
            categoria = cursor.fetchone()
            if not categoria:
                return None, 'Categoría no válida'
            
            cursor.execute("""
                INSERT INTO productos (
                    codigo, nombre, categoria_id, precio_compra, precio_venta,
                    stock_actual, stock_minimo, stock_maximo, descripcion, impuesto
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (*valores[:2], categoria['id'], *valores[2:]))
            conn.commit()
            
            new_id = cursor.lastrowid
            cursor.execute("""
                SELECT p.id, p.codigo, p.nombre, c.nombre as categoria, 
                       p.precio_compra, p.precio_venta, p.stock_actual, 
                       p.stock_minimo, p.stock_maximo, p.descripcion, p.impuesto
                FROM productos p
                JOIN categorias c ON p.categoria_id = c.id
                WHERE p.id = %s
            """, (new_id,))
            nuevo_producto = cursor.fetchone()
            
            nuevo_producto['codigo'] = nuevo_producto['codigo'] or f"PRD-{str(new_id).zfill(3)}"
            nuevo_producto['nombre'] = nuevo_producto['nombre'] or 'Producto sin nombre'
            nuevo_producto['categoria'] = nuevo_producto['categoria'] or 'Sin categoría'
            nuevo_producto['descripcion'] = nuevo_producto['descripcion'] or ''
            
            cursor.close()
            return {"message": "Producto creado exitosamente", "product": nuevo_producto}
        
        resultado, error = manejar_db(operacion)
        if resultado:
            return jsonify(resultado), 201
        else:
            return jsonify({'error': error}), 500

@bp.route('/api/productos/<int:id>', methods=['GET', 'PUT', 'DELETE'])
def handle_producto(id):
    if request.method == 'GET':
        def operacion(conn):
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT p.id, p.codigo, p.nombre, c.nombre as categoria, 
                       p.precio_compra, p.precio_venta, p.stock_actual, 
                       p.stock_minimo, p.stock_maximo, p.descripcion, p.impuesto
                FROM productos p
                JOIN categorias c ON p.categoria_id = c.id
                WHERE p.id = %s
            """, (id,))
            producto = cursor.fetchone()
            cursor.close()
            
            if not producto:
                return None, 'Producto no encontrado'
            
            producto['codigo'] = producto['codigo'] or 'SIN CÓDIGO'
            producto['nombre'] = producto['nombre'] or 'SIN NOMBRE'
            producto['categoria'] = producto['categoria'] or 'SIN CATEGORÍA'
            producto['descripcion'] = producto['descripcion'] or ''
            return producto
        
        resultado, error = manejar_db(operacion)
        if resultado:
            return jsonify(resultado)
        else:
            return jsonify({'error': error}), 404 if error == 'Producto no encontrado' else 500
    
    elif request.method == 'PUT':
        data = request.get_json()
        errores = validar_producto(data)
        if errores:
            return jsonify({'error': 'Errores de validación', 'detalles': errores}), 400
        
        try:
            valores = (
                data['code'],
                data['name'],
                float(data['purchase_price']),
                float(data['sale_price']),
                int(data['initial_quantity']),
                int(data['min_stock']),
                int(data['max_stock']),
                data.get('description', ''),
                float(data['tax_percentage']),
                id
            )
        except (TypeError, ValueError) as e:
            return jsonify({'error': f'Tipos de datos inválidos: {str(e)}'}), 400
        
        if valores[5] > valores[6]:
            return jsonify({'error': 'El stock mínimo no puede ser mayor que el stock máximo'}), 400
        
        def operacion(conn):
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT id FROM categorias WHERE nombre = %s", (data['category'],))
            categoria = cursor.fetchone()
            if not categoria:
                return None, 'Categoría no válida'
            
            cursor.execute("""
                UPDATE productos SET
                    codigo = %s,
                    nombre = %s,
                    categoria_id = %s,
                    precio_compra = %s,
                    precio_venta = %s,
                    stock_actual = %s,
                    stock_minimo = %s,
                    stock_maximo = %s,
                    descripcion = %s,
                    impuesto = %s
                WHERE id = %s
            """, (*valores[:2], categoria['id'], *valores[2:]))
            conn.commit()
            
            cursor.execute("""
                SELECT p.id, p.codigo, p.nombre, c.nombre as categoria, 
                       p.precio_compra, p.precio_venta, p.stock_actual, 
                       p.stock_minimo, p.stock_maximo, p.descripcion, p.impuesto
                FROM productos p
                JOIN categorias c ON p.categoria_id = c.id
                WHERE p.id = %s
            """, (id,))
            producto = cursor.fetchone()
            cursor.close()
            
            if not producto:
                return None, 'Producto no encontrado después de actualizar'
            
            producto['codigo'] = producto['codigo'] or 'SIN CÓDIGO'
            producto['nombre'] = producto['nombre'] or 'SIN NOMBRE'
            producto['categoria'] = producto['categoria'] or 'SIN CATEGORÍA'
            producto['descripcion'] = producto['descripcion'] or ''
            return producto
        
        resultado, error = manejar_db(operacion)
        if resultado:
            return jsonify(resultado)
        else:
            return jsonify({'error': error}), 500
    
    elif request.method == 'DELETE':
        def operacion(conn):
            cursor = conn.cursor()
            
            # Verificar si el producto existe
            cursor.execute("SELECT id FROM productos WHERE id = %s", (id,))
            if not cursor.fetchone():
                return None, "Producto no encontrado"
            
            # Eliminar movimientos y producto
            cursor.execute("DELETE FROM movimientos WHERE producto_id = %s", (id,))
            cursor.execute("DELETE FROM productos WHERE id = %s", (id,))
            conn.commit()
            cursor.close()
            return {'message': 'Producto eliminado exitosamente'}
        
        resultado, error = manejar_db(operacion)
        if resultado:
            return jsonify(resultado)
        else:
            status = 404 if "no encontrado" in error else 500
            return jsonify({'error': error}), status

@bp.route('/api/productos/proximo-codigo')
def get_proximo_codigo():
    def operacion(conn):
        cursor = conn.cursor()
        cursor.execute("SELECT MAX(codigo) FROM productos WHERE codigo LIKE 'PRD-%'")
        max_codigo = cursor.fetchone()[0]
        
        if max_codigo:
            try:
                num = int(max_codigo.split('-')[1]) + 1
            except (IndexError, ValueError):
                num = 1
        else:
            num = 1
            
        cursor.close()
        return {'proximo_codigo': f'PRD-{num:03d}'}
    
    resultado, error = manejar_db(operacion)
    if resultado:
        return jsonify(resultado)
    else:
        return jsonify({'error': error}), 500

# --- Nuevas funciones para importar/exportar Excel ---

@bp.route('/api/productos/exportar', methods=['GET'])
@login_required
@solo_admin_required
def exportar_productos():
    categoria_filtro = request.args.get('categoria', 'Todos')
    buscar_filtro = request.args.get('buscar', '').strip()
    
    def operacion(conn):
        productos = obtener_productos(conn, categoria_filtro, buscar_filtro)
        
        # Crear un libro de trabajo de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"
        
        # Encabezados
        encabezados = ['Código', 'Nombre', 'Descripción', 'Categoría', 'Precio Compra', 
                       'Precio Venta', 'Stock Actual', 'Stock Mínimo', 'Stock Máximo', 'Impuesto (%)']
        for col, encabezado in enumerate(encabezados, 1):
            ws.cell(row=1, column=col, value=encabezado)
        
        # Llenar con datos
        for row, producto in enumerate(productos, 2):
            ws.cell(row=row, column=1, value=producto.get('codigo', ''))
            ws.cell(row=row, column=2, value=producto.get('nombre', ''))
            ws.cell(row=row, column=3, value=producto.get('descripcion', ''))
            ws.cell(row=row, column=4, value=producto.get('categoria', ''))
            ws.cell(row=row, column=5, value=producto.get('precio_compra', 0))
            ws.cell(row=row, column=6, value=producto.get('precio_venta', 0))
            ws.cell(row=row, column=7, value=producto.get('stock_actual', 0))
            ws.cell(row=row, column=8, value=producto.get('stock_minimo', 0))
            ws.cell(row=row, column=9, value=producto.get('stock_maximo', 0))
            ws.cell(row=row, column=10, value=producto.get('impuesto', 0))
        
        # Ajustar el ancho de las columnas
        for col in range(1, len(encabezados) + 1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 15
        
        # Guardar en un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    resultado, error = manejar_db(operacion)
    if error:
        return jsonify({'error': error}), 500
    
    # Enviar el archivo
    fecha = datetime.datetime.now().strftime("%Y-%m-%d")
    return send_file(
        resultado,
        as_attachment=True,
        download_name=f'productos_{fecha}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@bp.route('/api/productos/importar', methods=['POST'])
@login_required
@solo_admin_required
def importar_productos():
    if 'file' not in request.files:
        return jsonify({'error': 'No se proporcionó ningún archivo'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No se seleccionó ningún archivo'}), 400
    
    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({'error': 'Formato de archivo no válido. Debe ser .xlsx o .xls'}), 400
    
    overwrite = request.form.get('overwrite', 'false').lower() == 'true'
    
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        # Leer encabezados
        headers = [cell.value for cell in ws[1]]
        
        # Mapeo de columnas (puede variar según el orden en el Excel)
        column_map = {}
        for idx, header in enumerate(headers, 1):
            if header == 'Código':
                column_map['codigo'] = idx
            elif header == 'Nombre':
                column_map['nombre'] = idx
            elif header == 'Descripción':
                column_map['descripcion'] = idx
            elif header == 'Categoría':
                column_map['categoria'] = idx
            elif header == 'Precio Compra':
                column_map['precio_compra'] = idx
            elif header == 'Precio Venta':
                column_map['precio_venta'] = idx
            elif header == 'Stock Actual':
                column_map['stock_actual'] = idx
            elif header == 'Stock Mínimo':
                column_map['stock_minimo'] = idx
            elif header == 'Stock Máximo':
                column_map['stock_maximo'] = idx
            elif header == 'Impuesto (%)':
                column_map['impuesto'] = idx
        
        # Validar que tenemos las columnas necesarias
        required_columns = ['nombre', 'categoria', 'precio_compra', 'precio_venta', 'stock_actual', 'stock_minimo', 'stock_maximo']
        for col in required_columns:
            if col not in column_map:
                return jsonify({'error': f'Falta la columna requerida: {col}'}), 400
        
        productos = []
        for row in range(2, ws.max_row + 1):
            # Leer cada fila
            producto = {
                'codigo': ws.cell(row=row, column=column_map.get('codigo', 0)).value,
                'nombre': ws.cell(row=row, column=column_map['nombre']).value,
                'descripcion': ws.cell(row=row, column=column_map.get('descripcion', 0)).value,
                'categoria_nombre': ws.cell(row=row, column=column_map['categoria']).value,
                'precio_compra': ws.cell(row=row, column=column_map['precio_compra']).value,
                'precio_venta': ws.cell(row=row, column=column_map['precio_venta']).value,
                'stock_actual': ws.cell(row=row, column=column_map['stock_actual']).value,
                'stock_minimo': ws.cell(row=row, column=column_map['stock_minimo']).value,
                'stock_maximo': ws.cell(row=row, column=column_map['stock_maximo']).value,
                'impuesto': ws.cell(row=row, column=column_map.get('impuesto', 0)).value or 0
            }
            
            # Validaciones básicas
            if not producto['nombre'] or not producto['categoria_nombre']:
                continue  # Saltar filas sin nombre o categoría
            
            # Convertir tipos numéricos
            try:
                producto['precio_compra'] = float(producto['precio_compra'] or 0)
                producto['precio_venta'] = float(producto['precio_venta'] or 0)
                producto['stock_actual'] = int(producto['stock_actual'] or 0)
                producto['stock_minimo'] = int(producto['stock_minimo'] or 0)
                producto['stock_maximo'] = int(producto['stock_maximo'] or 0)
                producto['impuesto'] = float(producto['impuesto'] or 0)
            except (ValueError, TypeError):
                # Si hay error en conversión, usar valores por defecto
                producto['precio_compra'] = 0
                producto['precio_venta'] = 0
                producto['stock_actual'] = 0
                producto['stock_minimo'] = 0
                producto['stock_maximo'] = 0
                producto['impuesto'] = 0
            
            productos.append(producto)
        
        # Ahora procesar los productos en la base de datos
        def operacion(conn):
            cursor = conn.cursor()
            errores = []
            insertados = 0
            actualizados = 0
            
            for producto in productos:
                # Verificar si la categoría existe, si no, crearla
                cursor.execute("SELECT id FROM categorias WHERE nombre = %s", (producto['categoria_nombre'],))
                categoria = cursor.fetchone()
                if not categoria:
                    # Crear la categoría
                    cursor.execute("INSERT INTO categorias (nombre) VALUES (%s)", (producto['categoria_nombre'],))
                    conn.commit()
                    categoria_id = cursor.lastrowid
                else:
                    categoria_id = categoria[0]
                
                # Verificar si el producto ya existe (por código o por nombre y categoría)
                if producto['codigo']:
                    cursor.execute("SELECT id FROM productos WHERE codigo = %s", (producto['codigo'],))
                    producto_existente = cursor.fetchone()
                else:
                    cursor.execute("SELECT id FROM productos WHERE nombre = %s AND categoria_id = %s", 
                                  (producto['nombre'], categoria_id))
                    producto_existente = cursor.fetchone()
                
                try:
                    if producto_existente and overwrite:
                        # Actualizar producto existente
                        cursor.execute("""
                            UPDATE productos SET
                                nombre = %s,
                                descripcion = %s,
                                categoria_id = %s,
                                precio_compra = %s,
                                precio_venta = %s,
                                stock_actual = %s,
                                stock_minimo = %s,
                                stock_maximo = %s,
                                impuesto = %s
                            WHERE id = %s
                        """, (
                            producto['nombre'],
                            producto['descripcion'],
                            categoria_id,
                            producto['precio_compra'],
                            producto['precio_venta'],
                            producto['stock_actual'],
                            producto['stock_minimo'],
                            producto['stock_maximo'],
                            producto['impuesto'],
                            producto_existente[0]
                        ))
                        actualizados += 1
                    else:
                        # Insertar nuevo producto
                        cursor.execute("""
                            INSERT INTO productos (
                                codigo, nombre, descripcion, categoria_id, 
                                precio_compra, precio_venta, stock_actual, 
                                stock_minimo, stock_maximo, impuesto
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            producto['codigo'],
                            producto['nombre'],
                            producto['descripcion'],
                            categoria_id,
                            producto['precio_compra'],
                            producto['precio_venta'],
                            producto['stock_actual'],
                            producto['stock_minimo'],
                            producto['stock_maximo'],
                            producto['impuesto']
                        ))
                        insertados += 1
                except Exception as e:
                    errores.append(f"Error con producto {producto['nombre']}: {str(e)}")
            
            conn.commit()
            cursor.close()
            
            return {
                'message': 'Importación completada',
                'insertados': insertados,
                'actualizados': actualizados,
                'errores': errores
            }
        
        resultado, error = manejar_db(operacion)
        if error:
            return jsonify({'error': error}), 500
        
        return jsonify(resultado)
        
    except Exception as e:
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}'}), 500

# Rutas para categorías
@bp.route('/api/categorias', methods=['GET'])
def listar_categorias():
    def operacion(conn):
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, nombre FROM categorias ORDER BY nombre")
        categorias = [{"id": r["id"], "nombre": r["nombre"]} for r in cursor.fetchall()]
        cursor.close()
        return categorias
    
    resultado, error = manejar_db(operacion)
    if resultado:
        return jsonify(resultado)
    else:
        return jsonify({"error": error}), 500

@bp.route('/api/categorias', methods=['POST'])
def crear_categoria():
    data = request.get_json() or {}
    nombre = (data.get('nombre') or '').strip()
    
    if not nombre:
        return jsonify({"error": "El nombre es requerido"}), 400
    
    def operacion(conn):
        cursor = conn.cursor()
        try:
            # Verificar si ya existe
            cursor.execute("SELECT id FROM categorias WHERE nombre = %s", (nombre,))
            if cursor.fetchone():
                return None, "Ya existe una categoría con ese nombre"
            
            cursor.execute("INSERT INTO categorias (nombre) VALUES (%s)", (nombre,))
            conn.commit()
            new_id = cursor.lastrowid
            return {"id": new_id, "nombre": nombre, "message": "Categoría creada exitosamente"}
        except Exception as e:
            conn.rollback()
            return None, f"Error de base de datos: {str(e)}"
        finally:
            cursor.close()
    
    resultado, error = manejar_db(operacion)
    if resultado:
        return jsonify(resultado), 201
    else:
        status = 409 if "Ya existe" in error else 500
        return jsonify({"error": error}), status

@bp.route('/api/categorias/<int:cat_id>', methods=['PUT'])
def actualizar_categoria(cat_id):
    data = request.get_json() or {}
    nombre = (data.get('nombre') or '').strip()
    if not nombre:
        return jsonify({"error": "El nombre es requerido"}), 400
    
    def operacion(conn):
        cursor = conn.cursor()
        try:
            # Verificar si la categoría existe
            cursor.execute("SELECT id FROM categorias WHERE id = %s", (cat_id,))
            if not cursor.fetchone():
                return None, "Categoría no encontrada"
            
            # Verificar si el nuevo nombre ya existe
            cursor.execute("SELECT id FROM categorias WHERE nombre = %s AND id != %s", (nombre, cat_id))
            if cursor.fetchone():
                return None, "Ya existe una categoría con ese nombre"
            
            cursor.execute("UPDATE categorias SET nombre = %s WHERE id = %s", (nombre, cat_id))
            conn.commit()
            return {"message": "Categoría actualizada exitosamente"}
        except Exception as e:
            conn.rollback()
            return None, f"Error de base de datos: {str(e)}"
        finally:
            cursor.close()
    
    resultado, error = manejar_db(operacion)
    if resultado:
        return jsonify(resultado)
    else:
        status = 404 if "no encontrada" in error else 409 if "Ya existe" in error else 500
        return jsonify({"error": error}), status

@bp.route('/api/categorias/<int:cat_id>', methods=['DELETE'])
def eliminar_categoria(cat_id):
    def operacion(conn):
        cursor = conn.cursor()
        try:
            # Verificar si la categoría existe
            cursor.execute("SELECT id FROM categorias WHERE id = %s", (cat_id,))
            if not cursor.fetchone():
                return None, "Categoría no encontrada"
            
            # Verificar si hay productos asociados
            cursor.execute("SELECT COUNT(*) FROM productos WHERE categoria_id = %s", (cat_id,))
            count = cursor.fetchone()[0]
            if count > 0:
                return None, "No se puede eliminar: la categoría tiene productos asociados"
            
            cursor.execute("DELETE FROM categorias WHERE id = %s", (cat_id,))
            conn.commit()
            return {"message": "Categoría eliminada exitosamente"}
        except Exception as e:
            conn.rollback()
            return None, f"Error de base de datos: {str(e)}"
        finally:
            cursor.close()
    
    resultado, error = manejar_db(operacion)
    if resultado:
        return jsonify(resultado)
    else:
        status = 404 if "no encontrada" in error else 409 if "No se puede eliminar" in error else 500
        return jsonify({"error": error}), status